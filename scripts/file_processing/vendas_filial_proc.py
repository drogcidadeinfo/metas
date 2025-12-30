import os
import glob
import gspread
import json
import time
import logging
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError
from openpyxl.styles import Font

# Config logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

def get_latest_file(directory='.', extension='xls'):
    files = glob.glob(os.path.join(directory, f'*.{extension}'))
    if not files:
        logging.warning("No files found with the specified extension.")
        return None
    return max(files, key=os.path.getmtime)

def retry_api_call(func, retries=3, delay=2):
    for i in range(retries):
        try:
            return func()
        except HttpError as error:
            if hasattr(error, "resp") and error.resp.status == 500:
                logging.warning(f"APIError 500 encountered. Retrying {i + 1}/{retries}...")
                time.sleep(delay)
            else:
                raise
    raise Exception("Max retries reached.")

def process_excel_data(file_path):
    logging.info("Processing Excel file (filial totals)...")

    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # ---- helper: dynamic column lookup ----
    def get_column_index(sheet, header_name):
        header_name = header_name.strip().lower()
        for col in sheet.iter_cols(1, sheet.max_column):
            cell_value = str(col[0].value).strip().lower()
            if cell_value == header_name:
                return col[0].column - 1
        raise ValueError(f"Cabeçalho '{header_name}' não encontrado")

    # ---- locate required columns ----
    col_codigo = get_column_index(sheet, 'código')
    col_total_vlr_venda = get_column_index(sheet, 'total vlr. venda')
    col_total_vlr_custo = get_column_index(sheet, 'total vlr. custo')
    col_vlr_descto = get_column_index(sheet, 'vlr. descto')
    col_ticket_medio = get_column_index(sheet, 'ticket médio venda/devol.')

    data = []

    filial_number = None
    faturamento_hb = None

    # ---- iterate rows ----
    for row in sheet.iter_rows(values_only=True):
        codigo = str(row[col_codigo]).strip() if row[col_codigo] else ""

        # Detect filial header
        if codigo.lower() == 'filial:':
            filial_number = str(row[col_codigo + 1]).split()[0]
            continue

        # Detect HB row
        if codigo == '8000':
            faturamento_hb = row[col_total_vlr_venda]
            continue

        # Detect total row
        if codigo.lower().startswith('total filial'):
            if not filial_number:
                logging.warning("Total Filial found without Filial number")
                continue

            data.append({
                'Filial': filial_number.zfill(2),
                'Faturamento HB': faturamento_hb,
                'Custo Total': row[col_total_vlr_custo],
                'Faturamento Total': row[col_vlr_descto],
                'Ticket Médio': row[col_ticket_medio],
            })

            # reset for next filial
            filial_number = None
            faturamento_hb = None

    result_df = pd.DataFrame(data)

    logging.info(f"Rows processed: {len(result_df)}")
    return result_df

def update_google_sheet(df, sheet_id, worksheet_name):
    logging.info("Updating Google Sheet...")

    creds_json = os.getenv("GSA_CREDENTIALS")
    creds = Credentials.from_service_account_info(
        json.loads(creds_json),
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
    )

    client = gspread.authorize(creds)
    worksheet = client.open_by_key(sheet_id).worksheet(worksheet_name)

    # Clear existing data
    worksheet.batch_clear(["A2:Z"])

    # Upload (headers + values)
    worksheet.update(
        range_name="A1",
        values=[df.columns.tolist()] + df.fillna("").values.tolist(),
        value_input_option="USER_ENTERED"
    )

    logging.info("Google Sheet updated successfully.")

def main():
    download_dir = "/home/runner/work/metas/metas/"
    sheet_id = os.getenv("SHEET_ID")

    time.sleep(15)

    file_path = get_latest_file(download_dir)

    if not file_path:
        logging.warning("No file found to process.")
        return

    logging.info(f"Processing file: {file_path}")

    try:
        df = process_excel_data(file_path)

        if df.empty:
            logging.warning("No valid rows found. Skipping upload.")
            return

        update_google_sheet(df, sheet_id, "VENDAS_FILIAL")

        os.remove(file_path)
        logging.info(f"File removed: {file_path}")

    except Exception as e:
        logging.error(f"Processing failed: {e}")

if __name__ == "__main__":
    main()
